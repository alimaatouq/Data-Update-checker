import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill
from io import BytesIO



def compare_and_save(original_file_path, edited_file_path):
    # Open the original and edited workbooks
    original_file = openpyxl.load_workbook(original_file_path)
    edited_file = openpyxl.load_workbook(edited_file_path)

    fill_style = PatternFill(start_color="FDD835", end_color="FDD835", fill_type="solid")

    # Get a list of sheet names to compare (assuming the sheet names are the same in both workbooks)
    sheet_names = original_file.sheetnames

    # Iterate through the sheet names
    for sheet_name in sheet_names:
        original_sheet = original_file[sheet_name]
        edited_sheet = edited_file[sheet_name]

        for row_original, row_edited in zip(original_sheet.iter_rows(), edited_sheet.iter_rows()):
            for cell_original, cell_edited in zip(row_original, row_edited):
                original_value = cell_original.value
                edited_value = cell_edited.value

                if original_value != edited_value:
                    cell_edited.fill = fill_style

    # Save the edited workbook to a BytesIO object
    output_file = BytesIO()
    edited_file.save(output_file)
    output_file.seek(0)

    return output_file.getvalue()

def main():
    st.image("https://www.google.com/url?sa=i&url=http%3A%2F%2Fwww.unescwa.org%2F&psig=AOvVaw0j8_PLBZzC7l52L5wEhqti&ust=1700221536719000&source=images&cd=vfe&ved=0CBIQjRxqFwoTCMCdvvi4yIIDFQAAAAAdAAAAABAT")
    st.title("Excel Comparison App")

    # Upload original and edited files
    original_file = st.file_uploader("Upload Original Excel File", type=["xlsx"])
    edited_file = st.file_uploader("Upload Edited Excel File", type=["xlsx"])

    if original_file and edited_file:
        # Perform comparison and get the compared file content
        compared_file_content = compare_and_save(original_file, edited_file)

        # Add a download button for the compared file
        st.download_button(
            label="Download Compared File",
            data=compared_file_content,
            file_name="compared_file.xlsx",
            key="download_button"
        )

if __name__ == "__main__":
    main()
#edit footer
page_style= """
    <style>
    footer{
        visibility: visible;
        }
    footer:after{
        content: 'Developed by Ali Maatouk - ESCWA Statistics, Information Society and Technology Cluster';
        display:block;
        position:relative;
        color:#1e54e4;
    }
    </style>"""

st.markdown(page_style, unsafe_allow_html=True)
